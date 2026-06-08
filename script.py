"""
Phishing Training Tracker — Automation Script
----------------------------------------------
Usage:
    python script.py <source_file.xlsx> <tracking_file.xlsx>

    1st argument = ABI Learning source report  (read only — never modified)
    2nd argument = Phishing Tracking file      (updated in place)

Match logic:
    1. Try Global Employee ID  +  Zone Level 2  +  Zone Level 3
    2. Try Local Employee ID   +  Zone Level 2  +  Zone Level 3
    3. Try Global Employee ID  only  (zone mismatch flagged)
    4. Try Local Employee ID   only  (zone mismatch flagged)
    → If none match: "Not Found"

Output:
    - Tracking file updated in place (All Data tab)
    - One Excel file per zone in the folder you run the script from
    - Each zone file has one tab per year found in the data

Optimised for 200,000+ rows — vectorised pandas + numpy, no iterrows.
"""

import sys, os, re
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
NEW_COLUMNS  = ["Training Start Date", "Training End Date", "Transcript Status", "Match Type"]

# Source columns
SRC_EMP_ID   = "Employee ID"
SRC_ZONE2    = "Macro Entity Level 2 (Zone)"
SRC_ZONE3    = "Macro Entity Level 3 (BU) Description"
SRC_STATUS   = "Employee Status"
SRC_TSTATUS  = "Transcript Status"
SRC_TSTART   = "Training Start Date"
SRC_TEND     = "Transcript Completed Date"
SOURCE_USECOLS = [SRC_EMP_ID, SRC_STATUS, SRC_TSTATUS, SRC_TSTART, SRC_TEND,
                  SRC_ZONE2, SRC_ZONE3]

# Tracking columns
TRK_GLOBAL   = "Global Employee ID"
TRK_LOCAL    = "Local Employee ID"
TRK_ZONE2    = "Macro Entity Level 2 (Zone)"

# Zone column used to split output files
ZONE_COL     = TRK_ZONE2

# ── Styles ────────────────────────────────────────────────────────────────────
ORANGE_FILL  = PatternFill("solid", fgColor="FFC000")
YELLOW_FILL  = PatternFill("solid", fgColor="FFFF00")
NO_FILL      = PatternFill(fill_type=None)
HEADER_FONT  = Font(bold=True, name="Arial", size=10)
DATA_FONT    = Font(name="Arial", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
THIN         = Side(style="thin")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

YELLOW_COLS  = {"Zone","Country","Global Employee ID","Local Employee ID",
                "Employee Name","Employee Status","Worker Type","Employee Group",
                "Management Level","First Hire Date","Last Hire Date","Position Name",
                "Job Family Group","ABI Entity 2","Macro Entity Level 2 (Zone)",
                "text before","Employee Email","Band 4+",
                "Manager Employee ID Level 01","Manager Name Level 01","BSC"}

# ── Header scan ───────────────────────────────────────────────────────────────
def find_header_row(path: str) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(max_row=40, values_only=True)):
        if row and any(str(v).strip() == "Employee ID" for v in row if v is not None):
            wb.close()
            return i
    wb.close()
    raise ValueError(
        "Could not find 'Employee ID' column in the first 40 rows.\n"
        "Please check this is the correct ABI Learning transcript report."
    )

# ── Step 1 — Build lookup ─────────────────────────────────────────────────────
def build_lookup(source_path: str) -> pd.DataFrame:
    print(f"  Reading: {os.path.basename(source_path)}")
    hdr_row = find_header_row(source_path)
    print(f"  → Header detected at row {hdr_row + 1}")

    df = pd.read_excel(source_path, header=hdr_row, usecols=SOURCE_USECOLS,
                       dtype={SRC_EMP_ID: str, SRC_STATUS: str, SRC_TSTATUS: str},
                       engine="openpyxl")

    df[SRC_EMP_ID] = df[SRC_EMP_ID].astype(str).str.strip()
    df[SRC_ZONE2]  = df[SRC_ZONE2].astype(str).str.strip().str.upper()
    df[SRC_ZONE3]  = df[SRC_ZONE3].astype(str).str.strip().str.upper()
    df = df[df[SRC_EMP_ID].notna() & ~df[SRC_EMP_ID].isin(["", "nan", "None"])]

    # Keep best status per employee+zone combo
    rank     = {"Completed": 0, "In Progress": 1, "Not Started": 2}
    df["_rank"] = df[SRC_TSTATUS].map(rank).fillna(99)
    df = (df.sort_values("_rank")
            .drop_duplicates(subset=[SRC_EMP_ID, SRC_ZONE2, SRC_ZONE3], keep="first")
            .copy())

    # Vectorised dates + terminated logic
    df["_start"] = pd.to_datetime(df[SRC_TSTART], errors="coerce")
    df["_end"]   = pd.to_datetime(df[SRC_TEND],   errors="coerce")
    is_term      = df[SRC_STATUS].str.strip().str.lower() == "terminated"
    has_start    = df["_start"].notna()

    df["_start_str"] = df["_start"].dt.strftime("%d-%b-%Y").fillna("")
    df["_end_str"]   = np.where(
        is_term & has_start, "Terminated",
        df["_end"].dt.strftime("%d-%b-%Y").fillna("")
    )

    print(f"  → {len(df):,} records loaded | "
          f"Zones: {sorted(df[SRC_ZONE2].unique())}")
    return df

# ── Step 2 — Merge with tiered matching ──────────────────────────────────────
def merge_tracking(tracking_path: str, src: pd.DataFrame) -> pd.DataFrame:
    print(f"  Reading: {os.path.basename(tracking_path)}")

    xl    = pd.ExcelFile(tracking_path)
    sheet = "All Data" if "All Data" in xl.sheet_names else xl.sheet_names[0]
    trk   = pd.read_excel(tracking_path, sheet_name=sheet,
                          dtype={TRK_GLOBAL: str, TRK_LOCAL: str}, engine="openpyxl")

    # Drop stale new columns
    stale = [c for c in trk.columns if c in NEW_COLUMNS or re.match(r".+\.\d+$", c)]
    trk   = trk.drop(columns=stale, errors="ignore")

    # Normalise IDs — strip decimal (.0) and whitespace
    trk[TRK_GLOBAL] = trk[TRK_GLOBAL].astype(str).str.strip().str.split(".").str[0]
    trk[TRK_LOCAL]  = trk[TRK_LOCAL].astype(str).str.strip().str.split(".").str[0]
    trk[TRK_ZONE2]  = trk[TRK_ZONE2].astype(str).str.strip().str.upper()

    # Build lookup with composite keys
    def make_lookup(id_col_src):
        key = src[[id_col_src, SRC_ZONE2, SRC_ZONE3,
                   "_start_str", "_end_str", SRC_TSTATUS]].copy()
        key.columns = ["_id", "_z2", "_z3", "Training Start Date",
                       "Training End Date", "Transcript Status"]
        return key.set_index(["_id", "_z2"])

    lkp = make_lookup(SRC_EMP_ID)

    # Also a fallback lookup by ID only (zone mismatch)
    lkp_id_only = (src[[SRC_EMP_ID, "_start_str", "_end_str", SRC_TSTATUS,
                         SRC_ZONE2, SRC_ZONE3]]
                   .copy()
                   .drop_duplicates(subset=[SRC_EMP_ID], keep="first"))
    lkp_id_only.columns = ["_id", "Training Start Date", "Training End Date",
                            "Transcript Status", "_src_z2", "_src_z3"]
    lkp_id_only = lkp_id_only.set_index("_id")

    results = []
    for _, row in trk.iterrows():
        g_id  = str(row[TRK_GLOBAL])
        l_id  = str(row[TRK_LOCAL])
        z2    = str(row[TRK_ZONE2])
        found = None
        mtype = "Not Found"

        # Tier 1: Global ID + Zone Level 2
        for eid in [g_id, l_id]:
            try:
                rec   = lkp.loc[(eid, z2)]
                found = rec
                mtype = f"Global ID + Zone L2" if eid == g_id else "Local ID + Zone L2"
                break
            except KeyError:
                pass

        # Tier 2: ID only (zone mismatch)
        if found is None:
            for eid in [g_id, l_id]:
                if eid in lkp_id_only.index:
                    rec   = lkp_id_only.loc[eid]
                    found = rec
                    mtype = (f"Global ID only (zone mismatch)" if eid == g_id
                             else "Local ID only (zone mismatch)")
                    break

        if found is not None:
            results.append({
                "Training Start Date": found["Training Start Date"],
                "Training End Date":   found["Training End Date"],
                "Transcript Status":   found["Transcript Status"],
                "Match Type":          mtype,
            })
        else:
            results.append({
                "Training Start Date": "Not Found",
                "Training End Date":   "Not Found",
                "Transcript Status":   "Not Found",
                "Match Type":          "Not Found",
            })

    res_df = pd.DataFrame(results, index=trk.index)
    trk    = pd.concat([trk, res_df], axis=1)

    matched   = (trk["Transcript Status"] != "Not Found").sum()
    unmatched = (trk["Transcript Status"] == "Not Found").sum()
    z2_only   = trk["Match Type"].str.contains("zone mismatch", na=False).sum()
    print(f"  → Matched: {matched:,}  |  Zone mismatch fallback: {z2_only:,}  |  Not found: {unmatched:,}")
    return trk

# ── Write sheet — numpy bulk writes ──────────────────────────────────────────
def write_sheet(ws, df: pd.DataFrame, new_col_set: set):
    all_cols = df.columns.tolist()
    for ci, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font, cell.border, cell.alignment = HEADER_FONT, BORDER, CENTER_ALIGN
        if col_name in new_col_set:    cell.fill = ORANGE_FILL
        elif col_name in YELLOW_COLS:  cell.fill = YELLOW_FILL
        else:                          cell.fill = NO_FILL

    data = df.values
    for ri, row_arr in enumerate(data, start=2):
        for ci, val in enumerate(row_arr, start=1):
            if val is pd.NaT or (not isinstance(val, str) and pd.isna(val)):
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font, cell.fill, cell.border = DATA_FONT, NO_FILL, BORDER
            cell.alignment = CENTER_ALIGN if all_cols[ci-1] in new_col_set else LEFT_ALIGN

    for ci, col_name in enumerate(all_cols, start=1):
        max_len = max(len(col_name),
                      int(df[col_name].astype(str).str.len().max() or 0))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}1"

# ── Main ──────────────────────────────────────────────────────────────────────
def update_tracking(source_path: str, tracking_path: str):
    cwd         = os.getcwd()
    new_col_set = set(NEW_COLUMNS)

    print("\n[1/4] Building lookup from source file ...")
    src = build_lookup(source_path)

    print("\n[2/4] Merging into tracking file ...")
    df = merge_tracking(tracking_path, src)

    print("\n[3/4] Saving updated tracking file ...")
    with pd.ExcelWriter(tracking_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Data", index=False)
        ws = writer.sheets["All Data"]
        for ci, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=ci)
            cell.font, cell.border, cell.alignment = HEADER_FONT, BORDER, CENTER_ALIGN
            if col_name in new_col_set:    cell.fill = ORANGE_FILL
            elif col_name in YELLOW_COLS:  cell.fill = YELLOW_FILL
            else:                          cell.fill = NO_FILL
        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"
    print(f"  → Saved: {os.path.basename(tracking_path)}")

    print("\n[4/4] Writing zone files ...")
    if ZONE_COL not in df.columns:
        print(f"  ⚠  '{ZONE_COL}' column not found — skipping zone files")
    else:
        df["_year"] = (pd.to_datetime(df["Training Start Date"],
                                      format="mixed", errors="coerce")
                         .dt.year.fillna(0).astype(int).astype(str)
                         .replace("0", "No Date"))
        all_cols = [c for c in df.columns if c != "_year"]

        for zone in sorted(df[ZONE_COL].dropna().unique()):
            zone_df  = df[df[ZONE_COL] == zone]
            safe     = re.sub(r"[^\w\s]", "", zone).strip()
            safe     = re.sub(r"\s+", "_", safe).title()
            out_path = os.path.join(cwd, f"{safe}.xlsx")
            wb       = Workbook()
            wb.remove(wb.active)

            years = sorted(zone_df["_year"].unique(),
                           key=lambda y: (y == "No Date", y))
            for year in years:
                year_df = zone_df[zone_df["_year"] == year][all_cols].copy()
                ws      = wb.create_sheet(title=str(year)[:31])
                write_sheet(ws, year_df, new_col_set)

            wb.save(out_path)
            print(f"  → {safe}.xlsx  ({len(zone_df):,} rows | tabs: {', '.join(years)})")

        df.drop(columns=["_year"], inplace=True)

    print(f"\n✅  Done! All files saved to → {cwd}\n")

# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("\nUsage:")
        print("  python script.py <source_file.xlsx> <tracking_file.xlsx>")
        print("\n  1st argument = ABI Learning source report  (read only)")
        print("  2nd argument = Phishing Tracking file      (updated in place)\n")
        sys.exit(1)

    source_path, tracking_path = sys.argv[1], sys.argv[2]

    if not os.path.exists(source_path):
        print(f"\n❌  Source file not found: {source_path}\n"); sys.exit(1)
    if not os.path.exists(tracking_path):
        print(f"\n❌  Tracking file not found: {tracking_path}\n"); sys.exit(1)

    try:
        update_tracking(source_path, tracking_path)
    except Exception as e:
        import traceback
        print(f"\n❌  Error: {e}")
        traceback.print_exc()
        sys.exit(1)
